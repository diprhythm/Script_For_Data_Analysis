# -*- coding: utf-8 -*-
"""
Excel 拆分高手
- 交互：多工作簿路径 → 选表 → 选拆分字段（支持 3 或 3+5）→ 输出目录
- 拆分：按“键值（字段/字段拼接）”→ 每个键一个工作簿；同键的不同源表写入该工作簿内的不同工作表
- 命名：工作表优先用“源 sheet 名”；若冲突自动加 _2/_3/...；不再加“文件名——”前缀
- 稳定：占位表 __INIT__（至少一个可见 Sheet）、原子保存、坏件自愈、规范化锁（绝对路径+小写）
- 并发：自动按 CPU，最大不超过 61；采用“在飞任务上限”限流 + wait(FIRST_COMPLETED) —— 杜绝 TimeoutError
- 日志：读前行数、写时每表行数、最终逐表校验（累计写出 == 源行数）
- 数字：读取 dtype=str；写出时智能列类型（金额/数量→数值；身份证/卡/账号/超长/前导零→文本）
依赖：
    pip install pandas openpyxl
"""

import os
import re
import sys
import time
import shutil
import tempfile
import traceback
import threading
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, wait, FIRST_COMPLETED

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# ---------- 日志 ----------
import logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)]
)
log = logging.getLogger("excel-splitter-pro-v2")

# ---------- 常量/规则 ----------
INVALID_WIN_CHARS = r'<>:"/\|?*'
INVALID_WIN_TRANS = str.maketrans({ch: "_" for ch in INVALID_WIN_CHARS})
SAFE_SHEET_MAX = 31
PLACEHOLDER_SHEET = "__INIT__"

ID_LIKE_KEYS = ["身份证", "证件", "卡", "账号", "账户", "手机", "电话", "流水号", "编号", "单号", "ID"]
NUMERIC_LIKE_KEYS = ["金额", "数", "余额", "收入", "支出", "单价", "数量", "合计", "价", "费", "税"]

# ---------- 工具 ----------
def sanitize_filename(name: str, maxlen: int = 200) -> str:
    name = (name or "").strip()
    if not name:
        name = "空值"
    name = name.translate(INVALID_WIN_TRANS)
    name = re.sub(r"[\r\n\t]+", " ", name)
    name = re.sub(r"\s+", " ", name).strip(" .")
    if not name:
        name = "空值"
    if len(name) > maxlen:
        name = name[:maxlen]
    return name

def sanitize_sheet_title(title: str) -> str:
    title = (title or "").translate(INVALID_WIN_TRANS)
    title = re.sub(r"[\[\]\:\*\?\/\\]", "_", title).strip()
    if len(title) > SAFE_SHEET_MAX:
        title = title[:SAFE_SHEET_MAX]
    return title or "Sheet"

def cpu_based_threads(max_cap: int = 61) -> int:
    c = os.cpu_count() or 4
    return min(max_cap, max(2, c * 2))

_num_re = re.compile(r"^[+-]?(\d+(\.\d+)?|\.\d+)$")

def looks_like_safe_number(s: str) -> bool:
    if s is None:
        return False
    s = str(s).strip()
    if not _num_re.match(s):
        return False
    if "." not in s:
        neg = s.startswith("-") or s.startswith("+")
        core = s[1:] if neg else s
        if len(core) > 1 and core.startswith("0"):
            return False
    digits_only = re.sub(r"[^0-9]", "", s)
    return len(digits_only) <= 15

def decide_column_is_numeric(col_name: str, series: pd.Series) -> bool:
    name = str(col_name)
    for k in ID_LIKE_KEYS:
        if k in name:
            return False
    hint = any(k in name for k in NUMERIC_LIKE_KEYS)
    non_empty = series.dropna().astype(str).str.strip()
    if non_empty.empty:
        return False
    safe_cnt = sum(looks_like_safe_number(v) for v in non_empty)
    ratio = safe_cnt / max(1, len(non_empty))
    if hint and ratio >= 0.5:
        return True
    if ratio >= 0.9:
        return True
    return False

def read_excel_headers(path: str, sheet_name: str) -> pd.Index:
    return pd.read_excel(path, sheet_name=sheet_name, dtype=str, nrows=0, engine="openpyxl").columns

def join_key_columns(df: pd.DataFrame, cols: list) -> pd.Series:
    if not cols:
        raise ValueError("用于拆分的字段未选择。")
    for c in cols:
        if c not in df.columns:
            raise ValueError(f"字段不存在：{c}")
    tmp = df[cols].astype(str).fillna("")
    key = tmp.apply(lambda row: " - ".join([v.strip() for v in row.values]), axis=1)
    key = key.apply(lambda s: s.strip() if isinstance(s, str) else s)
    return key.replace("", "空值")

# ---------- 文件锁（规范化） ----------
_file_locks = {}
_file_locks_guard = threading.Lock()

def _norm_lock_key(path: str) -> str:
    return os.path.abspath(path).lower()

def get_file_lock(path: str) -> threading.Lock:
    key = _norm_lock_key(path)
    with _file_locks_guard:
        if key not in _file_locks:
            _file_locks[key] = threading.Lock()
        return _file_locks[key]

# ---------- 原子保存 & 坏件自愈 ----------
def _safe_save_workbook(wb: Workbook, path: str, retries: int = 5) -> None:
    dir_ = os.path.dirname(path) or "."
    fd, tmp = tempfile.mkstemp(prefix="._writing_", suffix=".xlsx", dir=dir_)
    os.close(fd)
    last = None
    try:
        for i in range(retries):
            try:
                wb.save(tmp)
                os.replace(tmp, path)  # 原子替换
                last = None
                break
            except Exception as e:
                last = e
                time.sleep(0.2 * (2 ** i))  # 指数退避，抗杀软/占用
        if last is not None:
            raise last
    finally:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass

def ensure_workbook(path: str) -> None:
    p = Path(path)
    if not p.exists():
        wb = Workbook()
        wb.active.title = PLACEHOLDER_SHEET  # 保留占位可见表
        _safe_save_workbook(wb, str(p))

def _open_or_rebuild_workbook(path: str) -> Workbook:
    try:
        return load_workbook(path)
    except Exception as e:
        log.warning(f"检测到损坏的 xlsx，将重建：{path} | {e}")
        try:
            if os.path.exists(path):
                shutil.move(path, path + ".corrupt")
                log.warning(f"已备份损坏文件为：{path}.corrupt")
        except Exception:
            pass
        wb = Workbook()
        wb.active.title = PLACEHOLDER_SHEET
        _safe_save_workbook(wb, path)
        return load_workbook(path)

# ---------- 写入（智能类型 + 列宽 + 占位表清理 + 简洁命名） ----------
def _alloc_sheet_title(wb: Workbook, desired_title: str) -> str:
    """优先用 desired_title；若冲突，自动加 _2/_3/..."""
    base = sanitize_sheet_title(desired_title)
    if base not in wb.sheetnames:
        return base
    # 冲突时编号
    i = 2
    while True:
        cand = sanitize_sheet_title(f"{base}_{i}")
        if cand not in wb.sheetnames:
            return cand
        i += 1

def write_dataframe_as_new_sheet(xlsx_path: str, src_sheet_name: str, df: pd.DataFrame) -> int:
    """
    线程安全写入新工作表；返回写入的数据行数（不含表头）
    - 工作表名仅使用源 sheet 名；若冲突自动加 _2/_3/...
    - 首次写入真实表后，如有 __INIT__ 且总表数>1，则删除 __INIT__
    """
    lock = get_file_lock(xlsx_path)
    with lock:
        ensure_workbook(xlsx_path)
        wb = _open_or_rebuild_workbook(xlsx_path)

        title = _alloc_sheet_title(wb, src_sheet_name)
        ws = wb.create_sheet(title=title)

        # 列类型判定
        col_is_numeric = {col: decide_column_is_numeric(col, df[col]) for col in df.columns}

        # 表头
        headers = list(df.columns)
        ws.append(headers)

        # 数据先按文本写入
        records = df.replace({pd.NA: "", None: ""}).astype(str).values.tolist()
        for row in records:
            ws.append(row)

        # number_format & 转 float
        for j, col in enumerate(headers, start=1):
            col_letter = get_column_letter(j)
            if col_is_numeric[col]:
                num_fmt = "0.################"
                for cell in ws[col_letter][1:]:  # 跳过表头
                    s = str(cell.value).strip()
                    if looks_like_safe_number(s):
                        try:
                            cell.value = float(s)
                            cell.number_format = num_fmt
                        except Exception:
                            cell.number_format = "@"
                    else:
                        cell.number_format = "@"
            else:
                for cell in ws[col_letter][1:]:
                    cell.number_format = "@"

        # 简易列宽
        for j, col in enumerate(headers, start=1):
            sample = df[col].astype(str).tolist()
            max_len = max([len(str(col))] + [len(s) for s in sample[:200]])
            ws.column_dimensions[get_column_letter(j)].width = min(80, max(10, int(max_len * 1.2)))

        # 删除占位表（若已有真实表）
        if PLACEHOLDER_SHEET in wb.sheetnames and len(wb.sheetnames) > 1:
            try:
                wb.remove(wb[PLACEHOLDER_SHEET])
            except Exception:
                pass

        _safe_save_workbook(wb, xlsx_path)

    return len(df)

# ---------- 交互 ----------
def read_paths_from_stdin() -> list:
    print("请输入 Excel 文件完整路径（连续换行两次结束）：")
    lines, empty = [], 0
    while True:
        try:
            s = input().strip()
        except EOFError:
            break
        if s == "":
            empty += 1
            if empty >= 2:
                break
            continue
        empty = 0
        lines.append(s)
    uniq, valid = [], []
    for p in lines:
        if p not in uniq:
            uniq.append(p)
    for p in uniq:
        if Path(p).exists():
            valid.append(p)
        else:
            log.warning(f"文件不存在，已忽略：{p}")
    if not valid:
        raise SystemExit("未提供有效的 Excel 文件路径，退出。")
    return valid

def enumerate_all_sheets(paths: list) -> list:
    items, idx = [], 1
    print("\n已扫描到以下工作表：")
    for p in paths:
        try:
            xls = pd.ExcelFile(p, engine="openpyxl")
            sheets = xls.sheet_names
        except Exception as e:
            log.error(f"读取失败（跳过）：{p} | {e}")
            continue
        file_name = Path(p).name
        for s in sheets:
            display = f"{file_name}——{s}"
            items.append({"id": idx, "path": p, "file_name": file_name, "sheet_name": s, "display": display})
            print(f"{idx}：{display}")
            idx += 1
    if not items:
        raise SystemExit("未能读取到任何工作表。")
    return items

def parse_plus_numbers(s: str, upper: int) -> list:
    s = s.strip()
    if not s:
        return []
    parts = re.split(r"[+，,；;、\s]+", s)
    out = []
    for p in parts:
        if not p:
            continue
        if not p.isdigit():
            raise ValueError(f"非法编号：{p}")
        k = int(p)
        if k < 1 or k > upper:
            raise ValueError(f"编号超范围：{k}")
        if k not in out:
            out.append(k)
    return out

def interactive_select_tables(items: list) -> list:
    upper = items[-1]["id"]
    s = input("请输入需要参与拆分的表编号（如 1+3+4）：").strip()
    ids = parse_plus_numbers(s, upper)
    if not ids:
        raise SystemExit("未选择任何表，退出。")
    return [it for it in items if it["id"] in ids]

def interactive_select_fields_for_table(path: str, display: str, sheet_name: str) -> list:
    cols = read_excel_headers(path, sheet_name)
    print(f"\n请选择（{display}）中用于拆分的字段：")
    for i, c in enumerate(cols, start=1):
        print(f"{i}：{c}")
    while True:
        s = input("输入编号（支持 3 或 3+5）：").strip()
        try:
            ids = parse_plus_numbers(s, len(cols))
            if not ids:
                print("至少选择一个字段。")
                continue
            fields = [cols[i-1] for i in ids]
            print("已选字段：", " + ".join(fields))
            return fields
        except Exception as e:
            print(f"输入有误：{e}，请重试。")

# ---------- 主流程 ----------
def main():
    # 1) 输入文件路径
    paths = read_paths_from_stdin()

    # 2) 枚举所有表
    items = enumerate_all_sheets(paths)

    # 3) 选择要参与的表
    selected = interactive_select_tables(items)
    print("\n已选择的表：")
    for it in selected:
        print(f"- {it['display']}")

    # 4) 逐表选择拆分字段
    split_plan = []
    for it in selected:
        fields = interactive_select_fields_for_table(it["path"], it["display"], it["sheet_name"])
        split_plan.append({**it, "fields": fields})

    # 5) 输出目录
    out_dir = input("\n请输入拆分后文件的存放路径：").strip()
    if not out_dir:
        raise SystemExit("未提供输出路径。")
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    log.info(f"输出目录：{out_dir}")

    # 6) 读取每个源表、统计行数
    source_row_count = {}
    cached_full_df = {}
    for it in split_plan:
        try:
            df = pd.read_excel(it["path"], sheet_name=it["sheet_name"], dtype=str, engine="openpyxl")
        except Exception as e:
            log.error(f"读取失败（跳过）：{it['display']} | {e}")
            continue
        df = df.fillna("")
        cached_full_df[it["display"]] = df
        source_row_count[it["display"]] = len(df)
        log.info(f"[读取完成] {it['display']} | 行数：{len(df)}")

    if not source_row_count:
        raise SystemExit("没有可用的源表数据，退出。")

    # 7) 统计唯一键总数
    total_tasks = 0
    per_table_keys_count = []
    for it in split_plan:
        disp = it["display"]
        if disp not in cached_full_df:
            continue
        df = cached_full_df[disp]
        key = join_key_columns(df, it["fields"])
        n_unique = key.nunique(dropna=False)
        per_table_keys_count.append((it, n_unique))
        total_tasks += n_unique
        log.info(f"[预统计唯一键] {disp} | 唯一键：{n_unique}")
    if total_tasks == 0:
        raise SystemExit("没有可拆分的唯一键，退出。")

    threads = cpu_based_threads(max_cap=61)
    log.info(f"线程数：{threads}（自动估算，最大不超过 61）")

    # 8) 并发拆分写出（在飞任务上限）
    inflight_limit = threads * 4
    written_row_acc = {it["display"]: 0 for it in split_plan}
    done_counter = 0
    done_lock = threading.Lock()

    def task(xlsx_path: str, src_sheet_name: str, df_group: pd.DataFrame, key_value: str, src_display: str):
        nonlocal done_counter
        try:
            n = write_dataframe_as_new_sheet(xlsx_path, src_sheet_name, df_group)
            with done_lock:
                written_row_acc[src_display] = written_row_acc.get(src_display, 0) + n
                done_counter += 1
                log.info(f"[写出完成] {done_counter}/{total_tasks} | 键=[{key_value}] | "
                         f"→ 文件: {xlsx_path} | 表: {src_sheet_name} | 行数: {n}")
        except Exception:
            with done_lock:
                done_counter += 1
            log.error(f"[写出失败] 键=[{key_value}] | 源: {src_display}\n{traceback.format_exc()}")

    futures = set()
    start = time.time()
    with ThreadPoolExecutor(max_workers=threads, thread_name_prefix="split") as pool:
        for it, _n in per_table_keys_count:
            disp = it["display"]
            if disp not in cached_full_df:
                continue
            df = cached_full_df[disp].copy()
            key_series = join_key_columns(df, it["fields"])
            df["_SPLIT_KEY_"] = key_series
            for key_value, sub in df.groupby("_SPLIT_KEY_", dropna=False, sort=False):
                # 控制在飞上限
                while len(futures) >= inflight_limit:
                    done, futures = wait(futures, return_when=FIRST_COMPLETED)
                file_base = sanitize_filename(str(key_value))
                xlsx_path = os.path.join(out_dir, f"{file_base}.xlsx")
                fut = pool.submit(
                    task, xlsx_path, it["sheet_name"],
                    sub.drop(columns=["_SPLIT_KEY_"]), str(key_value), disp
                )
                futures.add(fut)
        # 收尾等待
        if futures:
            wait(futures)

    elapsed = time.time() - start
    log.info(f"[全部写出完成] 任务：{done_counter}/{total_tasks} | 耗时：{elapsed:.1f}s")

    # 9) 逐表校验
    log.info("[开始校验] 将拆分后各工作簿/表的行数相加，与拆分前源表行数对比：")
    all_ok = True
    for disp, src_rows in source_row_count.items():
        written_rows = written_row_acc.get(disp, 0)
        if written_rows == src_rows:
            log.info(f"[PASS] {disp} | 源行数={src_rows} | 写出合计={written_rows}")
        else:
            all_ok = False
            log.error(f"[FAIL] {disp} | 源行数={src_rows} | 写出合计={written_rows} | 请人工核查！")

    if all_ok:
        log.info(" 所有表格校验通过。")
    else:
        log.error(" 存在校验失败的表，请根据日志定位问题后复核。")

    input("校验完成，按回车结束进程...")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log.warning("用户中断。")
    except Exception as e:
        log.error(f"程序异常：{e}\n{traceback.format_exc()}")
        sys.exit(1)
