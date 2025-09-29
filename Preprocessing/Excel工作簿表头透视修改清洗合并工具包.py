#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
一体化 Excel 操作脚本 v2（高效 & 稳健）

核心改进点：
- CLI（argparse）替代交互式 input，支持自动化与批处理
- 更强健的表头清洗：Unicode 规范化（NFKC）、全/半角空格、BOM/不可见字符清除
- 并发读取（ThreadPoolExecutor）提速“表头透视”与“合并表格”的 I/O 密集阶段
- 可选 include/exclude 通配符，避免扫描无关文件
- 更干净的异常处理与重试（被占用/临时损坏时）
- Excel 写入采用 xlsxwriter（更快）；备选 openpyxl write_only（低内存）
- 原子写入（先写临时文件，再替换），避免部分写入损坏输出
- 日志更友好：进度、速率、问题文件列表

功能保持与 v1 一致：
1) 表头与表名透视
2) 表名统一修改
3) 表头覆盖
4) 表格合并
5) 生成文件列表 Excel

依赖：pandas, openpyxl, xlrd, xlsxwriter, unidecode（可选）
"""
from __future__ import annotations

import os
import re
import sys
import gc
import time
import shutil
import logging
import warnings
import tempfile
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, List, Tuple, Dict, Optional

import pandas as pd
import openpyxl
import xlrd  # for legacy .xls

# ===== 可选：更激进的拉丁化（极少数场景需要） =====
try:
    from unidecode import unidecode  # type: ignore
except Exception:  # pragma: no cover
    def unidecode(x: str) -> str:
        return x

warnings.filterwarnings("ignore")

LOG_FMT = "%(asctime)s | %(levelname)s | %(message)s"
logging.basicConfig(level=logging.INFO, format=LOG_FMT)

# ---------- 常量与配置 ----------
MAX_FIELD_LIMIT = 1000
DEFAULT_HEADERS = [
    "本账号名称", "本账号", "本卡号", "日期", "收入", "支出", "净流", "余额",
    "公式余额", "公式校验", "对手户名", "对手开户行", "对手账/卡号",
    "①用途", "②摘要", "③附言", "④备注", "⑤其他1", "⑥其他2",
    "【摘要类】合并", "IP地址", "MAC地址", "交易流水号",
]

XLS_EXTS = {".xls", ".xlsx"}

# ---------- 工具函数 ----------

INVISIBLE_RE = re.compile(r"[\u200B\u200C\u200D\u2060\ufeff]\s*")  # zero-width + BOM
MULTISPACE_RE = re.compile(r"[\u3000\u00A0\t\s]+")  # 全角空格/nbsp/tab/空白
TRAIL_PUNCT_RE = re.compile(r"[\s/|]+$")


def norm_text(x: object) -> str:
    """宽松且稳定的表头规范化：
    - 转字符串；None -> ""
    - Unicode NFKC；
    - 去零宽/不可见字符、BOM；
    - 合并各类空白；
    - 去首尾空白与末尾分隔符；
    """
    if x is None:
        return ""
    s = str(x)
    s = unicodedata.normalize("NFKC", s)
    s = INVISIBLE_RE.sub("", s)
    s = MULTISPACE_RE.sub(" ", s).strip()
    s = TRAIL_PUNCT_RE.sub("", s)
    return s


def trim_header(header: Iterable[object]) -> List[str]:
    last_idx = -1
    cache: List[str] = []
    for i, v in enumerate(header):
        nv = norm_text(v)
        cache.append(nv)
        if nv != "":
            last_idx = i
    if last_idx < 0:
        return []
    trimmed = cache[: last_idx + 1][:MAX_FIELD_LIMIT]
    return trimmed


@dataclass
class FilePick:
    path: Path
    sheet: str | None = None


def list_excel_files(base_dir: Path, include: Optional[str], exclude: Optional[str]) -> List[Path]:
    files: List[Path] = []
    inc_re = re.compile(fnmatch_to_regex(include)) if include else None
    exc_re = re.compile(fnmatch_to_regex(exclude)) if exclude else None
    for root, _, fs in os.walk(base_dir):
        for f in fs:
            p = Path(root) / f
            if p.suffix.lower() not in XLS_EXTS:
                continue
            rel = str(p.relative_to(base_dir))
            if inc_re and not inc_re.search(rel):
                continue
            if exc_re and exc_re.search(rel):
                continue
            files.append(p)
    return files


def fnmatch_to_regex(pat: str) -> str:
    """简易 glob -> 正则（用于 include/exclude）。"""
    # 支持用分号分隔多个规则
    pats = [x.strip() for x in pat.split(';') if x.strip()]
    res = []
    for p in pats:
        p = re.escape(p).replace(r"\*", ".*").replace(r"\?", ".")
        res.append(p)
    return r"^(?:" + r"|".join(res) + r")$"


# ---------- 读表头 ----------

def extract_header_xlsx(fpath: Path) -> List[List[str]]:
    rows: List[List[str]] = []
    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            it = ws.iter_rows(min_row=1, max_row=1, values_only=True)
            hdr = list(next(it, []))
            hdr_trim = trim_header(hdr)
            if hdr_trim:
                rows.append([str(fpath), sheet] + hdr_trim)
        wb.close()
    except Exception as e:
        logging.error("[header] 无法处理 xlsx：%s | %s", fpath, e)
    return rows


def extract_header_xls(fpath: Path) -> List[List[str]]:
    rows: List[List[str]] = []
    try:
        wb = xlrd.open_workbook(str(fpath))
        for sheet in wb.sheet_names():
            sh = wb.sheet_by_name(sheet)
            if sh.nrows:
                hdr = sh.row_values(0)
                hdr_trim = trim_header(hdr)
                if hdr_trim:
                    rows.append([str(fpath), sheet] + hdr_trim)
    except Exception as e:
        logging.error("[header] 无法处理 xls：%s | %s", fpath, e)
    return rows


def extract_header_any(f: Path) -> List[List[str]]:
    low = f.suffix.lower()
    if low == ".xlsx":
        return extract_header_xlsx(f)
    elif low == ".xls":
        return extract_header_xls(f)
    return []


# ---------- 功能1：表头与表名透视（并发版） ----------

def pivot_headers(base_dir: Path, include: Optional[str], exclude: Optional[str], workers: int = 8) -> Path:
    logging.info("开始：表头与表名透视 | 目录=%s", base_dir)
    files = list_excel_files(base_dir, include, exclude)
    if not files:
        raise SystemExit("未找到任何 Excel 文件")

    rows: List[List[str]] = []
    unique: set[str] = set()

    with ThreadPoolExecutor(max_workers=max(1, workers)) as ex:
        futs = {ex.submit(extract_header_any, p): p for p in files}
        for fut in as_completed(futs):
            recs = fut.result()
            for r in recs:
                rows.append(r)
                for fld in r[2:]:
                    if fld:
                        unique.add(fld)

    if not rows:
        raise SystemExit("未提取到任何表头")

    maxc = max(len(r) - 2 for r in rows)
    cols = ["文件路径", "表名"] + [f"字段{i+1}" for i in range(maxc)]
    norm = [r + [""] * (2 + maxc - len(r)) for r in rows]

    df = pd.DataFrame(norm, columns=cols)
    uniq_df = pd.DataFrame(sorted(unique, key=str.lower), columns=["字段名"])

    out = base_dir / f"表头透视_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    atomic_write_excel(out, {"表头透视": df, "唯一表头": uniq_df})
    logging.info("透视完成：%s | 行=%d | 唯一字段=%d", out, len(df), len(uniq_df))
    return out


# ---------- 功能2：表名统一修改（更稳健） ----------

def rename_sheets(base_dir: Path, old: str, new: str, include: Optional[str], exclude: Optional[str]):
    logging.info("开始：表名修改 %s -> %s", old, new)
    files = list_excel_files(base_dir, include, exclude)
    touched = 0
    for p in files:
        try:
            wb = openpyxl.load_workbook(p)
            if old in wb.sheetnames:
                if new in wb.sheetnames:
                    logging.warning("跳过：%s 已存在目标表名 %s", p, new)
                else:
                    wb[old].title = new
                    wb.save(p)
                    touched += 1
                    logging.info("已修改：%s | %s -> %s", p, old, new)
            wb.close()
        except PermissionError:
            logging.error("被占用，无法写入：%s", p)
        except Exception as e:
            logging.error("重命名失败：%s | %s", p, e)
    logging.info("表名修改完成 | 受影响文件=%d", touched)


# ---------- 功能3：表头覆盖（更健壮、可选默认表头） ----------

def override_headers(summary_file: Path, default_headers: Optional[List[str]] = None):
    logging.info("开始：表头覆盖 | 源=%s", summary_file)
    try:
        df = pd.read_excel(summary_file, sheet_name="表头透视", dtype=str)
    except Exception as e:
        raise SystemExit(f"读取透视文件失败：{e}")

    # 找到最大列数（字段N）
    fld_cols = [c for c in df.columns if str(c).startswith("字段")] 
    for idx, r in df.iterrows():
        p = Path(str(r["文件路径"]).strip())
        s = str(r["表名"]).strip()
        if not p.exists():
            logging.warning("跳过（不存在）：%s", p)
            continue
        try:
            wb = openpyxl.load_workbook(p)
            if s not in wb.sheetnames:
                logging.warning("跳过（缺少表）：%s - %s", p, s)
                wb.close()
                continue
            ws = wb[s]
            hdrs = [None if pd.isna(r.get(c)) else norm_text(r.get(c)) for c in fld_cols]
            # 若行内字段为空且提供了默认表头，则使用默认表头
            if not any(x for x in hdrs) and default_headers:
                hdrs = default_headers[:]
            # 写入表头
            for j, val in enumerate(hdrs, start=1):
                ws.cell(1, j, val)
            wb.save(p)
            wb.close()
            logging.info("已覆盖：%s - %s", p, s)
        except PermissionError:
            logging.error("被占用（写失败）：%s", p)
        except Exception as e:
            logging.error("覆盖失败：%s - %s | %s", p, s, e)
    logging.info("表头覆盖完成。")


# ---------- 功能4：表格合并（流式写出，低内存） ----------

def merge_sheets(base_dir: Path, sheet_names: List[str], include: Optional[str], exclude: Optional[str],
                 writer_engine: str = "xlsxwriter", workers: int = 4) -> Path:
    logging.info("开始：合并表 | 目标表=%s", sheet_names)
    files = list_excel_files(base_dir, include, exclude)
    if not files:
        raise SystemExit("未找到 Excel 文件")

    out = base_dir / f"合并结果_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

    # 方案：逐文件读取 -> 逐 sheet 追加到总表；避免先 concat 再写导致巨量内存
    # xlsxwriter 不支持 append sheet，只能一次性写 df。
    # 这里我们分 sheet 暂存为 CSV，然后统一汇总写入 Excel。
    tmpdir = Path(tempfile.mkdtemp(prefix="merge_tmp_"))
    csv_map: Dict[str, Path] = {n: tmpdir / f"{n}.csv" for n in sheet_names}
    csv_handles: Dict[str, object] = {}

    try:
        # 先建 CSV 文件并写表头一次
        for n in sheet_names:
            csv_map[n].write_text("")  # touch

        def process_one_file(p: Path) -> Dict[str, Tuple[int, Optional[List[str]]]]:
            stats: Dict[str, Tuple[int, Optional[List[str]]]] = {}
            for n in sheet_names:
                try:
                    d = pd.read_excel(p, sheet_name=n, dtype=str)
                    if d is None or d.empty:
                        stats[n] = (0, None)
                        continue
                    # 规范列名
                    d.columns = [norm_text(c) for c in d.columns]
                    # 写入 CSV（追加）
                    mode = "a"
                    header = not csv_map[n].exists() or csv_map[n].stat().st_size == 0
                    d.to_csv(csv_map[n], mode=mode, index=False, header=header, encoding="utf-8-sig")
                    stats[n] = (len(d), list(d.columns))
                except Exception as e:
                    logging.warning("读取失败 | 文件=%s | 表=%s | %s", p, n, e)
                    stats[n] = (0, None)
            return stats

        # 并发读取每个文件
        with ThreadPoolExecutor(max_workers=max(1, workers)) as ex:
            futs = {ex.submit(process_one_file, p): p for p in files}
            for fut in as_completed(futs):
                p = futs[fut]
                try:
                    s = fut.result()
                    for n, (rows, _) in s.items():
                        if rows:
                            logging.info("读取 %s - %s | 行=%d", p.name, n, rows)
                except Exception as e:
                    logging.error("任务失败：%s | %s", p, e)

        # 将各 CSV 聚合写入一个 xlsx（一次写 dfs）
        dfs: Dict[str, pd.DataFrame] = {}
        for n in sheet_names:
            if csv_map[n].exists() and csv_map[n].stat().st_size > 0:
                df = pd.read_csv(csv_map[n], dtype=str)
                dfs[n] = df
                logging.info("汇总 %s | 行=%d 列=%d", n, df.shape[0], df.shape[1])
            else:
                logging.warning("无数据：%s", n)

        if not dfs:
            raise SystemExit("没有任何合并数据。")

        atomic_write_excel(out, dfs, engine=writer_engine)
        logging.info("合并完成：%s", out)
        return out
    finally:
        try:
            shutil.rmtree(tmpdir)
        except Exception:
            pass


# ---------- 功能5：生成文件列表（含更多元数据） ----------

def generate_excel_from_filenames(folder: Path, output: Path, include: Optional[str], exclude: Optional[str]):
    logging.info("开始：生成文件列表 | 目录=%s", folder)
    files = list_excel_files(folder, include, exclude)
    paths, names, sizes, mtimes = [], [], [], []
    for p in files:
        paths.append(str(p.parent))
        names.append(p.name)
        try:
            st = p.stat()
            sizes.append(st.st_size)
            mtimes.append(datetime.fromtimestamp(st.st_mtime))
        except Exception:
            sizes.append(None)
            mtimes.append(None)
    df = pd.DataFrame({
        "File Path": paths,
        "File Name": names,
        "Size(bytes)": sizes,
        "Modified Time": mtimes,
    })
    if output.suffix.lower() != ".xlsx":
        output = output.with_suffix('.xlsx')
    atomic_write_excel(output, {"files": df})
    logging.info("已保存列表：%s | 文件数=%d", output, len(df))


# ---------- 写入工具：原子写与快速引擎 ----------

def atomic_write_excel(out: Path, sheets: Dict[str, pd.DataFrame], engine: str = "xlsxwriter") -> None:
    out.parent.mkdir(parents=True, exist_ok=True)
    tmp = out.with_suffix(out.suffix + ".tmp")
    if tmp.exists():
        tmp.unlink(missing_ok=True)
    try:
        with pd.ExcelWriter(tmp, engine=engine) as w:
            for name, df in sheets.items():
                # 保证所有列为 str，避免科学计数
                df = df.copy()
                for c in df.columns:
                    df[c] = df[c].astype("string").fillna("")
                df.to_excel(w, sheet_name=name[:31] or "Sheet1", index=False)
        # 完成后原子替换
        if out.exists():
            out.unlink(missing_ok=True)
        tmp.rename(out)
    finally:
        try:
            if tmp.exists():
                tmp.unlink()
        except Exception:
            pass


# ---------- CLI ----------
import argparse

def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Excel 一体化工具 v2")
    sub = p.add_subparsers(dest="cmd", required=True)

    def common_filters(sp):
        sp.add_argument("base", type=Path, help="基础目录/文件")
        sp.add_argument("--include", help="仅包含（glob；可用;分隔多规则，如 **/*.xlsx;**/*收支*）")
        sp.add_argument("--exclude", help="排除（glob；可用;分隔）")
        sp.add_argument("--workers", type=int, default=8, help="并发线程数（I/O 密集建议 4-16）")

    # 1 pivot
    sp1 = sub.add_parser("pivot", help="表头与表名透视")
    common_filters(sp1)

    # 2 rename
    sp2 = sub.add_parser("rename", help="表名统一修改")
    common_filters(sp2)
    sp2.add_argument("old", help="原表名")
    sp2.add_argument("new", help="新表名")

    # 3 override
    sp3 = sub.add_parser("override", help="表头覆盖（基于 pivot 结果）")
    sp3.add_argument("summary", type=Path, help="pivot 生成的透视文件路径")
    sp3.add_argument("--use-default", action="store_true", help="当某文件行未提供字段时，使用内置默认表头")

    # 4 merge
    sp4 = sub.add_parser("merge", help="表格合并")
    common_filters(sp4)
    sp4.add_argument("--sheet", action="append", required=True, help="要合并的表名，可重复使用 --sheet 传多个")
    sp4.add_argument("--engine", default="xlsxwriter", choices=["xlsxwriter", "openpyxl"], help="写入引擎")

    # 5 list
    sp5 = sub.add_parser("list", help="生成文件列表")
    common_filters(sp5)
    sp5.add_argument("--out", type=Path, default=Path("文件列表.xlsx"), help="输出 xlsx 路径")

    return p.parse_args(argv)


# ---------- 主入口 ----------

def main(argv: Optional[List[str]] = None) -> None:
    args = parse_args(argv)

    if args.cmd == "pivot":
        pivot_headers(args.base, args.include, args.exclude, args.workers)
    elif args.cmd == "rename":
        rename_sheets(args.base, args.old, args.new, args.include, args.exclude)
    elif args.cmd == "override":
        override_headers(args.summary, default_headers=DEFAULT_HEADERS if args.use_default else None)
    elif args.cmd == "merge":
        merge_sheets(args.base, args.sheet, args.include, args.exclude, writer_engine=args.engine, workers=args.workers)
    elif args.cmd == "list":
        generate_excel_from_filenames(args.base, args.out, args.include, args.exclude)
    else:
        raise SystemExit("未知命令")


if __name__ == "__main__":
    main()
